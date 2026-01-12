#!/usr/bin/env python3
"""
Generate Test Evaluation Report with Cosine Similarity

This script processes test inference results and generates an Excel report
with cosine similarity scores between predictions and ground truth.

Usage:
    python utils/generate_test_report.py --predictions test_predictions.json
    python utils/generate_test_report.py --predictions test_predictions.json --output test_report.xlsx
"""

import json
import argparse
import numpy as np
from pathlib import Path
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from sklearn.metrics.pairwise import cosine_similarity
import evaluate
from sentence_transformers import SentenceTransformer
import torch
from transformers import AutoTokenizer, AutoModelForCausalLM
import re


# ROUGE-L Score thresholds (0-1 scale, higher is better)
ROUGE_GREEN_THRESHOLD = 0.5   # >= this value is green (good)
ROUGE_YELLOW_THRESHOLD = 0.2  # >= this value is yellow (moderate), below is red

# BERT Similarity thresholds (0-1 scale, higher is better)
BERT_GREEN_THRESHOLD = 0.7   # >= this value is green (good)
BERT_YELLOW_THRESHOLD = 0.4  # >= this value is yellow (moderate), below is red

# METEOR Score thresholds (0-1 scale, higher is better)
METEOR_GREEN_THRESHOLD = 0.5   # >= this value is green (good)
METEOR_YELLOW_THRESHOLD = 0.2  # >= this value is yellow (moderate), below is red

# LLM Accuracy thresholds (1-5 scale, higher is better)
LLM_GREEN_THRESHOLD = 4.0   # >= this value is green (good)
LLM_YELLOW_THRESHOLD = 3.0  # >= this value is yellow (moderate), below is red


def compute_meteor_score(reference: str, hypothesis: str, metric) -> float:
    """Compute METEOR score."""
    if not reference or not hypothesis or metric is None:
        return 0.0

    try:
        return metric.compute(predictions=[hypothesis], references=[reference])['meteor']
    except:
        return 0.0



def compute_rouge_score(reference: str, hypothesis: str, metric) -> float:
    """Compute ROUGE-L score."""
    if not reference or not hypothesis or metric is None:
        return 0.0

    try:
        result = metric.compute(predictions=[hypothesis], references=[reference])
        return result['rougeL']
    except:
        return 0.0


def extract_exercise_name(text: str) -> str:
    """Extract exercise name from text (before the dash)."""
    if not text:
        return ""

    # Find first dash and get text before it
    if '-' in text:
        return text.split('-')[0].strip().lower()
    return text.strip().lower()


def check_exercise_match(ground_truth: str, prediction: str) -> bool:
    """Check if exercise names match between ground truth and prediction."""
    gt_exercise = extract_exercise_name(ground_truth)
    pred_exercise = extract_exercise_name(prediction)

    if not gt_exercise or not pred_exercise:
        return False

    return gt_exercise == pred_exercise


def load_llm_judge():
    """Load Mixtral-Instruct-0.1 model for LM-as-judge evaluation."""
    print("\nLoading LLM Judge (Mixtral-8x7B-Instruct-v0.1)...")
    print("This may take a few minutes on first run...")

    try:
        model_name = "mistralai/Mixtral-8x7B-Instruct-v0.1"

        # Load tokenizer
        tokenizer = AutoTokenizer.from_pretrained(model_name)

        # Load model with optimizations
        model = AutoModelForCausalLM.from_pretrained(
            model_name,
            torch_dtype=torch.float16,
            device_map="auto",
            load_in_8bit=True  # Use 8-bit quantization to reduce memory
        )

        print("✓ LLM Judge loaded successfully")
        return tokenizer, model

    except Exception as e:
        print(f"⚠ Warning: Could not load LLM judge: {e}")
        print("  LLM Accuracy scores will be skipped")
        return None, None


def compute_llm_accuracy_score(ground_truth: str, prediction: str, tokenizer, model) -> float:
    """
    Use Mixtral-Instruct as a judge to score prediction against ground truth.
    Returns a score from 1-5 for holistic accuracy and usefulness.

    Args:
        ground_truth: Reference feedback
        prediction: Model-generated feedback
        tokenizer: Mixtral tokenizer
        model: Mixtral model

    Returns:
        float: Score from 1.0 to 5.0
    """
    if not ground_truth or not prediction or tokenizer is None or model is None:
        return 0.0

    try:
        # Create prompt for LM-as-judge
        prompt = f"""[INST] You are an expert evaluator for exercise feedback quality.

Given a ground-truth feedback and a predicted feedback for a physiotherapy exercise video, rate the predicted feedback on a scale of 1-5 for holistic accuracy and usefulness.

Rating Scale:
5 - Excellent: Predicted feedback is highly accurate, covers all key points, and is very useful
4 - Good: Predicted feedback is mostly accurate with minor omissions, still quite useful
3 - Moderate: Predicted feedback has some accuracy but misses important details
2 - Poor: Predicted feedback has major inaccuracies or missing critical information
1 - Very Poor: Predicted feedback is largely incorrect or not useful

Ground-truth feedback:
{ground_truth}

Predicted feedback:
{prediction}

Provide only a single number (1-5) as your rating. [/INST]

Rating:"""

        # Tokenize
        inputs = tokenizer(prompt, return_tensors="pt", truncation=True, max_length=2048)
        inputs = {k: v.to(model.device) for k, v in inputs.items()}

        # Generate
        with torch.inference_mode():
            outputs = model.generate(
                **inputs,
                max_new_tokens=10,
                do_sample=False,
                temperature=0.0,
                pad_token_id=tokenizer.eos_token_id
            )

        # Decode response
        response = tokenizer.decode(outputs[0][inputs['input_ids'].shape[1]:], skip_special_tokens=True)

        # Extract score (look for first number 1-5)
        match = re.search(r'\b([1-5])\b', response)
        if match:
            score = float(match.group(1))
            return score
        else:
            print(f"⚠ Could not parse LLM response: {response[:100]}")
            return 3.0  # Default to middle score

    except Exception as e:
        print(f"⚠ Error computing LLM accuracy: {e}")
        return 0.0


def compute_cosine_similarity_bert(text1: str, text2: str, model) -> float:
    """Compute cosine similarity using BERT embeddings."""
    if not text1 or not text2:
        return 0.0

    try:
        embeddings = model.encode([text1, text2])
        similarity = cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]
        return float(similarity)
    except:
        return 0.0


def create_excel_report(results: List[Dict], output_path: str, use_bert: bool = True,
                       base_predictions: List[Dict] = None, use_llm_judge: bool = True):
    """Create an Excel report with formatted results and similarity scores."""

    # Create base prediction lookup if provided
    base_pred_map = {}
    if base_predictions:
        for bp in base_predictions:
            base_pred_map[bp.get('video', '')] = bp.get('base_prediction', '')
        print(f"✓ Loaded {len(base_pred_map)} base model predictions")

    # Load BERT model if requested
    bert_model = None
    if use_bert:
        print("Loading BERT model for semantic similarity...")
        try:
            bert_model = SentenceTransformer('all-MiniLM-L6-v2')
            print("✓ BERT model loaded")
        except Exception as e:
            print(f"⚠ Failed to load BERT model: {e}")
            print("  Falling back to TF-IDF similarity")
            use_bert = False

    # Load LLM judge if requested
    llm_tokenizer = None
    llm_model = None
    if use_llm_judge:
        llm_tokenizer, llm_model = load_llm_judge()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Evaluation Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "ID",
        "Video Path",
        "Ground Truth",
        "Model Prediction",
    ]

    # Add base model column if available
    if base_predictions:
        headers.append("Base Model Response")

    if use_bert:
        headers.append("BERT Similarity")

    headers.extend([
        "METEOR Score",
        "ROUGE-L Score",
        "LLM Accuracy (1-5)",
    ])

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Set column widths dynamically
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 40  # Video Path
    ws.column_dimensions['C'].width = 50  # Ground Truth
    ws.column_dimensions['D'].width = 50  # Prediction

    col_idx = 5  # Start after D

    if base_predictions:
        ws.column_dimensions[get_column_letter(col_idx)].width = 50  # Base Model Response
        col_idx += 1

    if use_bert:
        ws.column_dimensions[get_column_letter(col_idx)].width = 15  # BERT Similarity
        col_idx += 1

    ws.column_dimensions[get_column_letter(col_idx)].width = 15      # METEOR Score
    ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15  # ROUGE-L Score
    ws.column_dimensions[get_column_letter(col_idx + 2)].width = 18  # LLM Accuracy
    ws.column_dimensions[get_column_letter(col_idx + 3)].width = 15  # Throughput
    ws.column_dimensions[get_column_letter(col_idx + 4)].width = 12  # Gen Time
    ws.column_dimensions[get_column_letter(col_idx + 5)].width = 15  # Generated Tokens
    ws.column_dimensions[get_column_letter(col_idx + 6)].width = 15  # Exercise Match
    ws.column_dimensions[get_column_letter(col_idx + 7)].width = 10  # Status
    ws.column_dimensions[get_column_letter(col_idx + 8)].width = 30  # Error

    # Freeze header row
    ws.freeze_panes = "A2"

    # Process results
    print(f"\nProcessing {len(results)} results...")

    # Load METEOR metric
    meteor_metric = None
    try:
        meteor_metric = evaluate.load('meteor')
        print("✓ METEOR metric loaded")
    except Exception as e:
        print(f"⚠ Failed to load METEOR metric: {e}")

    # Load ROUGE metric
    rouge_metric = None
    try:
        rouge_metric = evaluate.load('rouge')
        print("✓ ROUGE metric loaded")
    except Exception as e:
        print(f"⚠ Failed to load ROUGE metric: {e}")

    bert_scores = []
    meteor_scores = []
    rouge_scores = []
    llm_accuracy_scores = []
    throughput_values = []
    generation_times = []
    exercise_matches = []
    exercise_stats = {}  # Track per-exercise statistics: {exercise_name: {'correct': x, 'total': y}}

    for idx, result in enumerate(results, start=1):
        row = idx + 1

        video_path = result.get('video_path', '')
        ground_truth = result.get('ground_truth', '')
        prediction = result.get('prediction', '')
        status = result.get('status', 'unknown')
        error = result.get('error', '')

        # Extract throughput metrics
        throughput = result.get('tokens_per_second', 0.0)
        gen_time = result.get('generation_time', 0.0)
        gen_tokens = result.get('generated_tokens', 0)

        if status == 'success' and throughput > 0:
            throughput_values.append(throughput)
            generation_times.append(gen_time)

        # Compute similarities
        bert_sim = None
        if use_bert and bert_model:
            bert_sim = compute_cosine_similarity_bert(ground_truth, prediction, bert_model)
            bert_scores.append(bert_sim)

        meteor_sim = compute_meteor_score(ground_truth, prediction, meteor_metric)
        meteor_scores.append(meteor_sim)

        rouge_sim = compute_rouge_score(ground_truth, prediction, rouge_metric)
        rouge_scores.append(rouge_sim)

        # Compute LLM accuracy (1-5 scale)
        llm_score = 0.0
        if use_llm_judge and llm_tokenizer and llm_model:
            llm_score = compute_llm_accuracy_score(ground_truth, prediction, llm_tokenizer, llm_model)
            llm_accuracy_scores.append(llm_score)

        exercise_match = check_exercise_match(ground_truth, prediction)
        exercise_matches.append(exercise_match)

        # Track per-exercise statistics
        exercise_name = extract_exercise_name(ground_truth)
        if exercise_name:
            if exercise_name not in exercise_stats:
                exercise_stats[exercise_name] = {'correct': 0, 'total': 0}
            exercise_stats[exercise_name]['total'] += 1
            if exercise_match:
                exercise_stats[exercise_name]['correct'] += 1

        # Write data
        col = 1
        ws.cell(row=row, column=col).value = idx
        col += 1
        ws.cell(row=row, column=col).value = video_path
        col += 1
        ws.cell(row=row, column=col).value = ground_truth
        col += 1
        ws.cell(row=row, column=col).value = prediction
        col += 1

        # Add base model response if available
        if base_predictions:
            base_pred = base_pred_map.get(video_path, 'N/A')
            ws.cell(row=row, column=col).value = base_pred
            col += 1

        # Track column indices for color coding
        bert_col_idx = None
        if use_bert and bert_sim is not None:
            ws.cell(row=row, column=col).value = round(bert_sim, 4)
            bert_col_idx = col
            col += 1

        meteor_col_idx = col
        ws.cell(row=row, column=col).value = round(meteor_sim, 4)
        col += 1

        rouge_col_idx = col
        ws.cell(row=row, column=col).value = round(rouge_sim, 4)
        col += 1

        # LLM Accuracy (1-5 scale)
        llm_col_idx = col
        if use_llm_judge and llm_score > 0:
            ws.cell(row=row, column=col).value = round(llm_score, 2)
        col += 1

        exercise_col_idx = col
        ws.cell(row=row, column=col).value = "TRUE" if exercise_match else "FALSE"
        col += 1

        ws.cell(row=row, column=col).value = status
        col += 1
        ws.cell(row=row, column=col).value = error

        # Apply formatting
        for c in range(1, col + 1):
            cell = ws.cell(row=row, column=c)
            cell.alignment = cell_alignment
            cell.border = border

            # Color code similarity scores
            if bert_col_idx and c == bert_col_idx:  # BERT column
                score = bert_sim if bert_sim is not None else 0
                if score >= BERT_GREEN_THRESHOLD:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= BERT_YELLOW_THRESHOLD:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            if c == meteor_col_idx:  # METEOR column
                score = meteor_sim
                if score >= METEOR_GREEN_THRESHOLD:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= METEOR_YELLOW_THRESHOLD:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            if c == rouge_col_idx:  # ROUGE-L column
                score = rouge_sim
                if score >= ROUGE_GREEN_THRESHOLD:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= ROUGE_YELLOW_THRESHOLD:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            if use_llm_judge and c == llm_col_idx:  # LLM Accuracy column
                score = llm_score
                if score >= LLM_GREEN_THRESHOLD:  # 4-5: Excellent/Good
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= LLM_YELLOW_THRESHOLD:  # 3: Moderate
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif score > 0:  # 1-2: Poor/Very Poor
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            if c == exercise_col_idx:  # Exercise Identified column
                if exercise_match:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")

    summary_data = [
        ["Metric", "Value"],
        ["Total Samples", len(results)],
        ["Successful", sum(1 for r in results if r.get('status') == 'success')],
        ["Failed", sum(1 for r in results if r.get('status') == 'error')],
        ["", ""],
    ]

    # Add throughput statistics
    if throughput_values:
        summary_data.extend([
            ["Throughput (tokens/sec)", ""],
            ["Mean", round(np.mean(throughput_values), 2)],
            ["Median", round(np.median(throughput_values), 2)],
            ["Std Dev", round(np.std(throughput_values), 2)],
            ["Min", round(np.min(throughput_values), 2)],
            ["Max", round(np.max(throughput_values), 2)],
            ["", ""],
        ])

    if generation_times:
        summary_data.extend([
            ["Generation Time (seconds)", ""],
            ["Mean", round(np.mean(generation_times), 4)],
            ["Median", round(np.median(generation_times), 4)],
            ["Std Dev", round(np.std(generation_times), 4)],
            ["Min", round(np.min(generation_times), 4)],
            ["Max", round(np.max(generation_times), 4)],
            ["", ""],
        ])

    # Track chart data positions
    bert_chart_start_row = None
    meteor_chart_start_row = None

    if use_bert and bert_scores:
        bert_green = sum(1 for s in bert_scores if s >= BERT_GREEN_THRESHOLD)
        bert_yellow = sum(1 for s in bert_scores if BERT_YELLOW_THRESHOLD <= s < BERT_GREEN_THRESHOLD)
        bert_red = sum(1 for s in bert_scores if s < BERT_YELLOW_THRESHOLD)
        bert_chart_start_row = len(summary_data) + 7  # Row where Green count will be
        summary_data.extend([
            ["BERT Similarity", ""],
            ["Mean", round(np.mean(bert_scores), 4)],
            ["Median", round(np.median(bert_scores), 4)],
            ["Std Dev", round(np.std(bert_scores), 4)],
            ["Min", round(np.min(bert_scores), 4)],
            ["Max", round(np.max(bert_scores), 4)],
            [f"Green (≥{BERT_GREEN_THRESHOLD})", bert_green],
            [f"Yellow ({BERT_YELLOW_THRESHOLD}-{BERT_GREEN_THRESHOLD})", bert_yellow],
            [f"Red (<{BERT_YELLOW_THRESHOLD})", bert_red],
            ["", ""],
        ])

    if meteor_scores:
        meteor_green = sum(1 for s in meteor_scores if s >= METEOR_GREEN_THRESHOLD)
        meteor_yellow = sum(1 for s in meteor_scores if METEOR_YELLOW_THRESHOLD <= s < METEOR_GREEN_THRESHOLD)
        meteor_red = sum(1 for s in meteor_scores if s < METEOR_YELLOW_THRESHOLD)
        meteor_chart_start_row = len(summary_data) + 7  # Row where Green count will be
        summary_data.extend([
            ["METEOR Score", ""],
            ["Mean", round(np.mean(meteor_scores), 4)],
            ["Median", round(np.median(meteor_scores), 4)],
            ["Std Dev", round(np.std(meteor_scores), 4)],
            ["Min", round(np.min(meteor_scores), 4)],
            ["Max", round(np.max(meteor_scores), 4)],
            [f"Green (≥{METEOR_GREEN_THRESHOLD})", meteor_green],
            [f"Yellow ({METEOR_YELLOW_THRESHOLD}-{METEOR_GREEN_THRESHOLD})", meteor_yellow],
            [f"Red (<{METEOR_YELLOW_THRESHOLD})", meteor_red],
            ["", ""],
        ])

    rouge_chart_start_row = None
    if rouge_scores:
        rouge_green = sum(1 for s in rouge_scores if s >= ROUGE_GREEN_THRESHOLD)
        rouge_yellow = sum(1 for s in rouge_scores if ROUGE_YELLOW_THRESHOLD <= s < ROUGE_GREEN_THRESHOLD)
        rouge_red = sum(1 for s in rouge_scores if s < ROUGE_YELLOW_THRESHOLD)
        rouge_chart_start_row = len(summary_data) + 7
        summary_data.extend([
            ["ROUGE-L Score", ""],
            ["Mean", round(np.mean(rouge_scores), 4)],
            ["Median", round(np.median(rouge_scores), 4)],
            ["Std Dev", round(np.std(rouge_scores), 4)],
            ["Min", round(np.min(rouge_scores), 4)],
            ["Max", round(np.max(rouge_scores), 4)],
            [f"Green (≥{ROUGE_GREEN_THRESHOLD})", rouge_green],
            [f"Yellow ({ROUGE_YELLOW_THRESHOLD}-{ROUGE_GREEN_THRESHOLD})", rouge_yellow],
            [f"Red (<{ROUGE_YELLOW_THRESHOLD})", rouge_red],
            ["", ""],
        ])

    if llm_accuracy_scores:
        llm_green = sum(1 for s in llm_accuracy_scores if s >= LLM_GREEN_THRESHOLD)
        llm_yellow = sum(1 for s in llm_accuracy_scores if LLM_YELLOW_THRESHOLD <= s < LLM_GREEN_THRESHOLD)
        llm_red = sum(1 for s in llm_accuracy_scores if s > 0 and s < LLM_YELLOW_THRESHOLD)
        summary_data.extend([
            ["LLM Accuracy (1-5)", ""],
            ["Mean", round(np.mean(llm_accuracy_scores), 2)],
            ["Median", round(np.median(llm_accuracy_scores), 2)],
            ["Std Dev", round(np.std(llm_accuracy_scores), 2)],
            ["Min", round(np.min(llm_accuracy_scores), 2)],
            ["Max", round(np.max(llm_accuracy_scores), 2)],
            [f"Green (≥{LLM_GREEN_THRESHOLD})", llm_green],
            [f"Yellow ({LLM_YELLOW_THRESHOLD}-{LLM_GREEN_THRESHOLD})", llm_yellow],
            [f"Red (<{LLM_YELLOW_THRESHOLD})", llm_red],
            ["", ""],
        ])

    if exercise_matches:
        exercise_correct = sum(1 for match in exercise_matches if match)
        exercise_incorrect = sum(1 for match in exercise_matches if not match)
        exercise_accuracy = (exercise_correct / len(exercise_matches) * 100) if exercise_matches else 0
        summary_data.extend([
            ["Exercise Identification", ""],
            ["Overall Correct", exercise_correct],
            ["Overall Incorrect", exercise_incorrect],
            ["Overall Accuracy (%)", round(exercise_accuracy, 2)],
            ["", ""],
            ["Per-Exercise Breakdown", ""],
        ])

        # Add per-exercise statistics sorted by exercise name
        for exercise_name in sorted(exercise_stats.keys()):
            stats = exercise_stats[exercise_name]
            correct = stats['correct']
            total = stats['total']
            # Format: "exercise_name" -> "x/total"
            display_name = exercise_name.replace('_', ' ').title()
            summary_data.append([display_name, f"{correct}/{total}"])

        summary_data.append(["", ""])

    for row_idx, row_data in enumerate(summary_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            if row_idx == 1 or (isinstance(value, str) and value and row_idx > 1 and col_idx == 1):
                cell.font = Font(bold=True)

    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 15

    if use_bert and bert_scores and bert_chart_start_row:
        # BERT Similarity Chart
        bert_chart = BarChart()
        bert_chart.type = "col"
        bert_chart.style = 10
        bert_chart.title = "BERT Similarity Distribution"
        bert_chart.y_axis.title = "Count"
        bert_chart.x_axis.title = "Category"

        # Show primary axes
        bert_chart.x_axis.delete = False
        bert_chart.y_axis.delete = False
        bert_chart.x_axis.majorTickMark = "out"
        bert_chart.y_axis.majorTickMark = "out"

        # Legend position below x-axis
        bert_chart.legend = Legend()
        bert_chart.legend.position = "b"

        # Data reference (values)
        bert_data = Reference(summary_ws, min_col=2, min_row=bert_chart_start_row,
                              max_row=bert_chart_start_row + 2)
        # Categories reference (labels)
        bert_cats = Reference(summary_ws, min_col=1, min_row=bert_chart_start_row,
                              max_row=bert_chart_start_row + 2)

        bert_chart.add_data(bert_data, titles_from_data=False)
        bert_chart.set_categories(bert_cats)
        bert_chart.shape = 4
        bert_chart.width = 12
        bert_chart.height = 8

        # Color the bars: green, yellow, red
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        from openpyxl.chart.shapes import GraphicalProperties

        series = bert_chart.series[0]
        # Green bar
        pt_green = DataPoint(idx=0)
        pt_green.graphicalProperties = GraphicalProperties()
        pt_green.graphicalProperties.solidFill = "00B050"
        series.data_points.append(pt_green)
        # Yellow bar
        pt_yellow = DataPoint(idx=1)
        pt_yellow.graphicalProperties = GraphicalProperties()
        pt_yellow.graphicalProperties.solidFill = "FFC000"
        series.data_points.append(pt_yellow)
        # Red bar
        pt_red = DataPoint(idx=2)
        pt_red.graphicalProperties = GraphicalProperties()
        pt_red.graphicalProperties.solidFill = "FF0000"
        series.data_points.append(pt_red)

        summary_ws.add_chart(bert_chart, "D2")

    if meteor_scores and meteor_chart_start_row:
        # METEOR Score Chart
        meteor_chart = BarChart()
        meteor_chart.type = "col"
        meteor_chart.style = 10
        meteor_chart.title = "METEOR Score Distribution"
        meteor_chart.y_axis.title = "Count"
        meteor_chart.x_axis.title = "Category"

        # Show primary axes
        meteor_chart.x_axis.delete = False
        meteor_chart.y_axis.delete = False
        meteor_chart.x_axis.majorTickMark = "out"
        meteor_chart.y_axis.majorTickMark = "out"

        # Legend position below x-axis
        meteor_chart.legend = Legend()
        meteor_chart.legend.position = "b"

        # Data reference (values)
        meteor_data = Reference(summary_ws, min_col=2, min_row=meteor_chart_start_row,
                                max_row=meteor_chart_start_row + 2)
        # Categories reference (labels)
        meteor_cats = Reference(summary_ws, min_col=1, min_row=meteor_chart_start_row,
                                max_row=meteor_chart_start_row + 2)

        meteor_chart.add_data(meteor_data, titles_from_data=False)
        meteor_chart.set_categories(meteor_cats)
        meteor_chart.shape = 4
        meteor_chart.width = 12
        meteor_chart.height = 8

        # Color the bars: green, yellow, red
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties

        series = meteor_chart.series[0]
        # Green bar
        pt_green = DataPoint(idx=0)
        pt_green.graphicalProperties = GraphicalProperties()
        pt_green.graphicalProperties.solidFill = "00B050"
        series.data_points.append(pt_green)
        # Yellow bar
        pt_yellow = DataPoint(idx=1)
        pt_yellow.graphicalProperties = GraphicalProperties()
        pt_yellow.graphicalProperties.solidFill = "FFC000"
        series.data_points.append(pt_yellow)
        # Red bar
        pt_red = DataPoint(idx=2)
        pt_red.graphicalProperties = GraphicalProperties()
        pt_red.graphicalProperties.solidFill = "FF0000"
        series.data_points.append(pt_red)

        # Position second chart below the first
        chart_position = "D17" if use_bert else "D2"
        summary_ws.add_chart(meteor_chart, chart_position)

    if rouge_scores and rouge_chart_start_row:
        # ROUGE-L Score Chart
        rouge_chart = BarChart()
        rouge_chart.type = "col"
        rouge_chart.style = 10
        rouge_chart.title = "ROUGE-L Score Distribution"
        rouge_chart.y_axis.title = "Count"
        rouge_chart.x_axis.title = "Category"

        # Show primary axes
        rouge_chart.x_axis.delete = False
        rouge_chart.y_axis.delete = False
        rouge_chart.x_axis.majorTickMark = "out"
        rouge_chart.y_axis.majorTickMark = "out"

        # Legend position below x-axis
        rouge_chart.legend = Legend()
        rouge_chart.legend.position = "b"

        # Data reference (values)
        rouge_data = Reference(summary_ws, min_col=2, min_row=rouge_chart_start_row,
                                max_row=rouge_chart_start_row + 2)
        # Categories reference (labels)
        rouge_cats = Reference(summary_ws, min_col=1, min_row=rouge_chart_start_row,
                                max_row=rouge_chart_start_row + 2)

        rouge_chart.add_data(rouge_data, titles_from_data=False)
        rouge_chart.set_categories(rouge_cats)
        rouge_chart.shape = 4
        rouge_chart.width = 12
        rouge_chart.height = 8

        # Color the bars: green, yellow, red
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties

        series = rouge_chart.series[0]
        # Green bar
        pt_green = DataPoint(idx=0)
        pt_green.graphicalProperties = GraphicalProperties()
        pt_green.graphicalProperties.solidFill = "00B050"
        series.data_points.append(pt_green)
        # Yellow bar
        pt_yellow = DataPoint(idx=1)
        pt_yellow.graphicalProperties = GraphicalProperties()
        pt_yellow.graphicalProperties.solidFill = "FFC000"
        series.data_points.append(pt_yellow)
        # Red bar
        pt_red = DataPoint(idx=2)
        pt_red.graphicalProperties = GraphicalProperties()
        pt_red.graphicalProperties.solidFill = "FF0000"
        series.data_points.append(pt_red)

        # Position third chart
        if use_bert and meteor_scores:
            chart_position = "D32"  # Below METEOR chart
        elif meteor_scores:
            chart_position = "D17"  # Below METEOR chart
        else:
            chart_position = "D2"   # First chart
        summary_ws.add_chart(rouge_chart, chart_position)

    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel report saved to: {output_path}")

    # Print summary
    print(f"\n{'='*60}")
    print("Evaluation Summary")
    print(f"{'='*60}")
    print(f"Total samples: {len(results)}")
    print(f"Successful: {sum(1 for r in results if r.get('status') == 'success')}")
    print(f"Failed: {sum(1 for r in results if r.get('status') == 'error')}")

    if throughput_values:
        print(f"\nThroughput (tokens/second):")
        print(f"  Mean: {np.mean(throughput_values):.2f}")
        print(f"  Median: {np.median(throughput_values):.2f}")
        print(f"  Std Dev: {np.std(throughput_values):.2f}")

    if generation_times:
        print(f"\nGeneration Time (seconds):")
        print(f"  Mean: {np.mean(generation_times):.4f}")
        print(f"  Median: {np.median(generation_times):.4f}")
        print(f"  Std Dev: {np.std(generation_times):.4f}")

    if use_bert and bert_scores:
        print(f"\nBERT Similarity:")
        print(f"  Mean: {np.mean(bert_scores):.4f}")
        print(f"  Median: {np.median(bert_scores):.4f}")
        print(f"  Std Dev: {np.std(bert_scores):.4f}")

    if meteor_scores:
        print(f"\nMETEOR Score:")
        print(f"  Mean: {np.mean(meteor_scores):.4f}")
        print(f"  Median: {np.median(meteor_scores):.4f}")
        print(f"  Std Dev: {np.std(meteor_scores):.4f}")

    if rouge_scores:
        print(f"\nROUGE-L Score:")
        print(f"  Mean: {np.mean(rouge_scores):.4f}")
        print(f"  Median: {np.median(rouge_scores):.4f}")
        print(f"  Std Dev: {np.std(rouge_scores):.4f}")

    if llm_accuracy_scores:
        print(f"\nLLM Accuracy (1-5 scale):")
        print(f"  Mean: {np.mean(llm_accuracy_scores):.2f}")
        print(f"  Median: {np.median(llm_accuracy_scores):.2f}")
        print(f"  Std Dev: {np.std(llm_accuracy_scores):.2f}")
        print(f"  Accuracy: {exercise_accuracy:.2f}%")

    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description="Generate test evaluation report with similarity scores")
    parser.add_argument("--predictions", type=str, required=True,
                        help="Path to predictions JSON from test_inference.py")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel file path (default: same directory as predictions)")
    parser.add_argument("--no-bert", action="store_true",
                        help="Skip BERT similarity (faster, uses only TF-IDF)")
    parser.add_argument("--no-llm-judge", action="store_true",
                        help="Skip LLM judge evaluation (faster, skips Mixtral scoring)")
    parser.add_argument("--include-base-model", action="store_true",
                        help="Include base model predictions (requires user confirmation)")
    parser.add_argument("--base-model", type=str, default="Amshaker/Mobile-VideoGPT-0.5B",
                        help="Base model path/ID for comparison")
    parser.add_argument("--test-json", type=str,
                        help="Test JSON file for base model inference")
    parser.add_argument("--data-path", type=str, default="dataset",
                        help="Base path for video files")

    args = parser.parse_args()

    # Set default output path to same directory as predictions if not provided
    if args.output is None:
        pred_path = Path(args.predictions)
        args.output = str(pred_path.parent / "test_evaluation_report.xlsx")
        print(f"Output will be saved to: {args.output}")

    # Load predictions
    print(f"Loading predictions from: {args.predictions}")
    with open(args.predictions, 'r') as f:
        results = json.load(f)

    print(f"Loaded {len(results)} predictions")

    # Handle base model predictions
    base_predictions = None
    if args.include_base_model:
        print("\n" + "="*60)
        print("BASE MODEL INFERENCE")
        print("="*60)
        print("⚠ Warning: This will run inference on all test samples")
        print("  using the base (non-finetuned) model.")
        print("  This may take significant time and compute resources.")
        print("="*60)

        response = input("\nProceed with base model inference? (yes/no): ").strip().lower()

        if response in ['yes', 'y', 'ok']:
            # Import base model inference utility
            try:
                import torch
                from utils.base_model_inference import get_base_model_predictions

                # Load test data
                if not args.test_json:
                    print("❌ Error: --test-json required for base model inference")
                    return

                with open(args.test_json, 'r') as f:
                    test_data = json.load(f)

                # Run base model inference
                base_predictions = get_base_model_predictions(
                    test_data,
                    base_model=args.base_model,
                    data_path=args.data_path,
                    device="cuda" if torch.cuda.is_available() else "cpu"
                )

                print("✓ Base model predictions obtained")

            except Exception as e:
                print(f"❌ Error during base model inference: {e}")
                print("  Continuing without base model predictions...")
                base_predictions = None
        else:
            print("Skipping base model inference.")

    # Generate report
    create_excel_report(results, args.output, use_bert=not args.no_bert,
                       base_predictions=base_predictions, use_llm_judge=not args.no_llm_judge)
if __name__ == "__main__":
    main()